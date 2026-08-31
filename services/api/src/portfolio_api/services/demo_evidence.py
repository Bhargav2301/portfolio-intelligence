from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

MAX_CLAIMS = 100
MARKET_METRIC_NAMES = {
    "allocation",
    "change",
    "current_price",
    "gain",
    "gain_percent",
    "high",
    "investment_amount",
    "investment_price",
    "latest_price",
    "latest_value",
    "low",
    "market_cap",
    "open",
    "overall_gain",
    "overall_gain_percent",
    "pb",
    "pe",
    "price",
    "quantity",
    "return_1_day",
    "return_1_month",
    "return_1_week",
    "return_1_year",
    "return_3_months",
    "return_6_months",
    "score",
    "stock_score",
    "target",
    "target_price",
    "upside",
    "upside_percent",
    "value",
}


@dataclass(frozen=True)
class DemoEvidenceExtraction:
    source_type: str
    title: str
    publisher: str
    published_at: datetime
    claims: tuple[dict[str, Any], ...]
    instruments: tuple[str, ...]
    parser_name: str
    parser_version: str = "1.0.0"

    def summary(self) -> dict[str, Any]:
        return {
            "parser_name": self.parser_name,
            "parser_version": self.parser_version,
            "claim_count": len(self.claims),
            "instruments": list(self.instruments),
            "source_type": self.source_type,
            "publisher": self.publisher,
            "published_at": self.published_at.isoformat(),
        }


def _slug(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9]+", "_", value.upper()).strip("_")
    return normalized[:64] or "DOCUMENT"


def _metric(value: str) -> str:
    raw = value.lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", raw).strip("_")
    if "%" in raw and normalized in {"gain", "overall_gain", "upside"}:
        normalized += "_percent"
    aliases = {
        "1d_return": "return_1_day",
        "1_day_return": "return_1_day",
        "1w_return": "return_1_week",
        "1_week_return": "return_1_week",
        "1m_return": "return_1_month",
        "1_month_return": "return_1_month",
        "3m_return": "return_3_months",
        "3_month_return": "return_3_months",
        "6m_return": "return_6_months",
        "6_month_return": "return_6_months",
        "1y_return": "return_1_year",
        "1_year_return": "return_1_year",
        "inv_amt": "investment_amount",
        "inv_price": "investment_price",
        "overall_gain": "overall_gain",
        "overall_gain_percent": "overall_gain_percent",
    }
    return aliases.get(normalized, normalized)[:64]


def _decimal(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    normalized = str(value).strip().replace(",", "").replace("₹", "").replace("Rs.", "")
    normalized = normalized.removesuffix("%").strip()
    if not normalized or normalized.lower() in {"n.a.", "na", "nan", "none", "-"}:
        return None
    try:
        parsed = Decimal(normalized)
    except InvalidOperation:
        return None
    return parsed if parsed.is_finite() else None


def _number_text(value: Decimal) -> str:
    return format(value.normalize(), "f")


def _claim(
    instrument: str,
    metric: str,
    statement: str,
    value: Decimal | None = None,
    unit: str | None = None,
) -> dict[str, Any]:
    return {
        "claim_key": f"instrument.{_slug(instrument)}.{_metric(metric)}",
        "statement": statement[:500],
        "numeric_value": _number_text(value) if value is not None else None,
        "unit": unit,
    }


def _parse_date(text: str) -> datetime | None:
    match = re.search(
        r"Last Updated On:\s*(\d{1,2}\s+[A-Za-z]{3},\s*\d{4},\s*\d{1,2}:\d{2}\s*[AP]M)\s*IST",
        text,
        re.IGNORECASE,
    )
    if match:
        parsed = datetime.strptime(re.sub(r"\s+", " ", match.group(1)), "%d %b, %Y, %I:%M %p")
        return parsed.replace(tzinfo=ZoneInfo("Asia/Kolkata"))
    match = re.search(r"(?:August|Aug)\s+(\d{1,2}),\s*(\d{4})", text, re.IGNORECASE)
    if match:
        return datetime(
            int(match.group(2)), 8, int(match.group(1)), tzinfo=ZoneInfo("Asia/Kolkata")
        )
    return None


def _add_numeric_match(
    claims: list[dict[str, Any]],
    text: str,
    pattern: str,
    *,
    instrument: str,
    metric: str,
    unit: str,
    label: str,
    flags: int = re.IGNORECASE,
) -> None:
    match = re.search(pattern, text, flags)
    value = _decimal(match.group(1)) if match else None
    if value is not None:
        claims.append(
            _claim(
                instrument, metric, f"{label} for {instrument} in the uploaded source.", value, unit
            )
        )


def extract_research_text(filename: str, text: str) -> DemoEvidenceExtraction | None:
    normalized_text = "\n".join(line.strip() for line in text.replace("\u00a0", " ").splitlines())
    now = datetime.now(UTC)
    published_at = _parse_date(normalized_text) or now
    if published_at > now:
        published_at = now
    if "CONTRACT NOTE CUM TAX INVOICE" in normalized_text and "As per Annexure" in normalized_text:
        date_match = re.search(
            r"TRADE DATE.*?(\d{2}-[A-Z]{3}-\d{4})",
            normalized_text,
            re.IGNORECASE | re.DOTALL,
        )
        if date_match:
            published_at = datetime.strptime(date_match.group(1).upper(), "%d-%b-%Y").replace(
                tzinfo=ZoneInfo("Asia/Kolkata")
            )
        trade_pattern = re.compile(
            r"As per Annexure\s+(?P<name>[A-Z][A-Z0-9 &.]+?)-\s*"
            r"(?P<isin>INE[A-Z0-9]{9})\s*(?P<side>Buy|Sell)\s+"
            r"(?P<quantity>[\d,]+(?:\.\d+)?)\s+"
            r"(?P<gross_rate>[\d,]+(?:\.\d+)?)\s+"
            r"(?P<brokerage_rate>[\d,]+(?:\.\d+)?)\s+"
            r"(?P<net_rate>[\d,]+(?:\.\d+)?)\s+"
            r"(?P<net_total>[\d,]+(?:\.\d+)?)",
            re.IGNORECASE,
        )
        claims: list[dict[str, Any]] = []
        instruments: list[str] = []
        for match in trade_pattern.finditer(normalized_text):
            instrument = re.sub(r"\s+", " ", match.group("name")).strip()
            instruments.append(instrument)
            claims.append(
                _claim(
                    instrument,
                    "trade_side",
                    f"The uploaded contract note records a {match.group('side').lower()} trade for {instrument}.",
                )
            )
            for metric, group, unit, label in (
                ("quantity", "quantity", "units", "Contract-note quantity"),
                ("execution_price", "gross_rate", "INR", "Contract-note gross rate"),
                ("brokerage_rate", "brokerage_rate", "INR", "Contract-note brokerage per unit"),
                ("net_rate", "net_rate", "INR", "Contract-note net rate per unit"),
                ("net_total", "net_total", "INR", "Contract-note net total before levies"),
            ):
                value = _decimal(match.group(group))
                if value is not None:
                    claims.append(
                        _claim(instrument, metric, f"{label} for {instrument}.", value, unit)
                    )
        if claims:
            return DemoEvidenceExtraction(
                source_type="market",
                title=f"Uploaded contract note - {Path(filename).stem}"[:255],
                publisher="Julius Baer Wealth Advisors (India)",
                published_at=published_at,
                claims=tuple(claims[:MAX_CLAIMS]),
                instruments=tuple(dict.fromkeys(instruments)),
                parser_name="spi-julius-baer-contract-note-pdf",
            )
    if "Economic Times" in normalized_text or "ETPrime" in normalized_text:
        instrument = Path(filename).stem.split(" Share Price", 1)[0].strip()
        claims: list[dict[str, Any]] = []
        for metric, label in (("open", "Open price"), ("high", "High price"), ("low", "Low price")):
            _add_numeric_match(
                claims,
                normalized_text,
                rf"{label.split()[0]}:\s*([\d,]+(?:\.\d+)?)",
                instrument=instrument,
                metric=metric,
                unit="INR",
                label=label,
            )
        _add_numeric_match(
            claims,
            normalized_text,
            r"share price is Rs\s*([\d,]+(?:\.\d+)?)",
            instrument=instrument,
            metric="latest_price",
            unit="INR",
            label="Latest reported share price",
        )
        returns_block = normalized_text
        block_match = re.search(
            r"Share Price\s*Returns(?P<body>.*?)(?:The Numbers Behind|ET Stock Screeners|News & Analysis)",
            normalized_text,
            re.IGNORECASE | re.DOTALL,
        )
        if block_match:
            returns_block = block_match.group("body")
        return_patterns = (
            ("return_1_day", r"1 Day[^\d-]*(-?\d+(?:\.\d+)?)%", "One-day return"),
            ("return_1_week", r"1 Week[^\d-]*(-?\d+(?:\.\d+)?)%", "One-week return"),
            ("return_1_month", r"1 Month[^\d-]*(-?\d+(?:\.\d+)?)%", "One-month return"),
            ("return_3_months", r"3 Months[^\d-]*(-?\d+(?:\.\d+)?)%", "Three-month return"),
            ("return_6_months", r"6 Months[^\d-]*(-?\d+(?:\.\d+)?)%", "Six-month return"),
            ("return_1_year", r"1 Year[^\d-]*(-?\d+(?:\.\d+)?)%", "One-year return"),
        )
        for metric, pattern, label in return_patterns:
            _add_numeric_match(
                claims,
                returns_block,
                pattern,
                instrument=instrument,
                metric=metric,
                unit="percent",
                label=label,
            )
        _add_numeric_match(
            claims,
            normalized_text,
            r"Stock\s*Score\s*(\d+(?:\.\d+)?)",
            instrument=instrument,
            metric="stock_score",
            unit="score_10",
            label="Publisher stock score",
        )
        outlook = re.search(r"\b(POSITIVE|NEGATIVE|NEUTRAL)\s+Outlook\b", normalized_text)
        if outlook:
            claims.append(
                _claim(
                    instrument,
                    "outlook",
                    f"The uploaded publisher page labels {instrument} with a {outlook.group(1).lower()} outlook.",
                )
            )
        _add_numeric_match(
            claims,
            normalized_text,
            r"PE ratio[^.]{0,80}?stands at\s*([\d.]+)",
            instrument=instrument,
            metric="pe",
            unit="ratio",
            label="Price-to-earnings ratio",
        )
        _add_numeric_match(
            claims,
            normalized_text,
            r"PB ratio(?:\s+is)?\s*([\d.]+)",
            instrument=instrument,
            metric="pb",
            unit="ratio",
            label="Price-to-book ratio",
        )
        if not claims:
            return None
        return DemoEvidenceExtraction(
            source_type="market",
            title=f"Uploaded market snapshot - {instrument}"[:255],
            publisher="The Economic Times",
            published_at=published_at,
            claims=tuple(claims[:MAX_CLAIMS]),
            instruments=(instrument,),
            parser_name="spi-economic-times-pdf",
        )

    if "Watchlist Gainers" in normalized_text and "Watchlist Losers" in normalized_text:
        row_pattern = re.compile(
            r"^(?P<name>[A-Za-z][A-Za-z0-9& .'-]+?)\s+"
            r"(?P<price>[\d,]+(?:\.\d+)?)\s+[^\d-]*(?P<day>-?\d+(?:\.\d+)?)%\s+"
            r"[^\d-]*(?P<price_change>-?\d+(?:\.\d+)?)\s+"
            r"[^\d-]*(?P<week>-?\d+(?:\.\d+)?)%\s+"
            r"[^\d-]*(?P<month>-?\d+(?:\.\d+)?)%\s+"
            r"[^\d-]*(?P<three_month>-?\d+(?:\.\d+)?)%\s+"
            r"[^\d-]*(?P<six_month>-?\d+(?:\.\d+)?)%$",
            re.MULTILINE,
        )
        claims = []
        instruments: list[str] = []
        for match in row_pattern.finditer(normalized_text):
            instrument = match.group("name").strip()
            instruments.append(instrument)
            for metric, group, unit, label in (
                ("latest_price", "price", "INR", "Watchlist current price"),
                ("return_1_day", "day", "percent", "Watchlist one-day return"),
                ("return_1_week", "week", "percent", "Watchlist one-week return"),
                ("return_1_month", "month", "percent", "Watchlist one-month return"),
                ("return_3_months", "three_month", "percent", "Watchlist three-month return"),
                ("return_6_months", "six_month", "percent", "Watchlist six-month return"),
            ):
                value = _decimal(match.group(group))
                if value is not None:
                    claims.append(
                        _claim(instrument, metric, f"{label} for {instrument}.", value, unit)
                    )
        if claims:
            return DemoEvidenceExtraction(
                source_type="market",
                title="Uploaded watchlist market snapshot",
                publisher="ET Markets",
                published_at=published_at,
                claims=tuple(claims[:MAX_CLAIMS]),
                instruments=tuple(dict.fromkeys(instruments)),
                parser_name="spi-watchlist-xray-pdf",
            )
    return None


def _pdf_text(content: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content), strict=True)
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _sheet_rows(filename: str, content: bytes) -> list[list[object]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        try:
            for name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(name)
                if sheet.nrows:
                    return [sheet.row_values(index) for index in range(sheet.nrows)]
        finally:
            workbook.release_resources()
        return []
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if rows:
                return rows
    finally:
        workbook.close()
    return []


def extract_research_rows(filename: str, rows: list[list[object]]) -> DemoEvidenceExtraction | None:
    if not rows:
        return None
    headers = [_metric(str(value or "")) for value in rows[0]]
    header_index = {name: index for index, name in enumerate(headers) if name}
    now = datetime.now(UTC)
    if {"stock", "latest_price", "quantity", "investment_amount", "latest_value"}.issubset(
        header_index
    ):
        aggregate: dict[str, dict[str, Decimal]] = {}
        display_names: dict[str, str] = {}
        for row in rows[1:]:
            stock_cell = row[header_index["stock"]] if header_index["stock"] < len(row) else None
            if not stock_cell:
                continue
            instrument = str(stock_cell).split(" - ", 1)[0].strip()
            key = _slug(instrument)
            if key == "TOTAL":
                continue
            display_names[key] = instrument
            bucket = aggregate.setdefault(
                key,
                {
                    "quantity": Decimal(0),
                    "investment_amount": Decimal(0),
                    "latest_value": Decimal(0),
                    "overall_gain": Decimal(0),
                    "latest_price": Decimal(0),
                },
            )
            for metric in tuple(bucket):
                index = header_index.get(metric)
                value = _decimal(row[index]) if index is not None and index < len(row) else None
                if value is None:
                    continue
                if metric == "latest_price":
                    bucket[metric] = value
                else:
                    bucket[metric] += value
        claims: list[dict[str, Any]] = []
        for key, values in aggregate.items():
            instrument = display_names[key]
            quantity = values["quantity"]
            investment_amount = values["investment_amount"]
            overall_gain = values["overall_gain"]
            derived = {
                **values,
                "investment_price": (investment_amount / quantity if quantity else Decimal(0)),
                "overall_gain_percent": (
                    overall_gain / investment_amount * Decimal(100)
                    if investment_amount
                    else Decimal(0)
                ),
            }
            for metric, unit, label in (
                ("quantity", "units", "Aggregated quantity"),
                ("investment_price", "INR", "Weighted investment price"),
                ("latest_price", "INR", "Latest workbook price"),
                ("latest_value", "INR", "Aggregated latest value"),
                ("overall_gain_percent", "percent", "Derived overall gain percentage"),
            ):
                claims.append(
                    _claim(
                        instrument,
                        metric,
                        f"{label} for {instrument} in the uploaded workbook.",
                        derived[metric],
                        unit,
                    )
                )
        return DemoEvidenceExtraction(
            source_type="market",
            title=f"Uploaded portfolio market snapshot - {Path(filename).stem}"[:255],
            publisher="User-provided workbook",
            published_at=now,
            claims=tuple(claims[:MAX_CLAIMS]),
            instruments=tuple(display_names.values()),
            parser_name="spi-holdings-workbook",
        )

    instrument_column = next(
        (name for name in ("instrument", "symbol", "stock", "company") if name in header_index),
        None,
    )
    claims = []
    instruments: list[str] = []
    for row_number, row in enumerate(rows[1:], start=2):
        instrument = (
            str(row[header_index[instrument_column]]).strip()
            if instrument_column and header_index[instrument_column] < len(row)
            else f"ROW_{row_number}"
        )
        if instrument and instrument != f"ROW_{row_number}":
            instruments.append(instrument)
        for header, index in header_index.items():
            if (
                header == instrument_column
                or header not in MARKET_METRIC_NAMES
                or index >= len(row)
            ):
                continue
            value = _decimal(row[index])
            if value is None:
                continue
            unit = "percent" if "percent" in header or "return" in header else "INR"
            if header == "quantity" or header == "score" or header == "stock_score":
                unit = "units" if header == "quantity" else "score_10"
            claims.append(
                _claim(
                    instrument,
                    header,
                    f"Uploaded tabular metric {header} for {instrument}.",
                    value,
                    unit,
                )
            )
            if len(claims) >= MAX_CLAIMS:
                break
        if len(claims) >= MAX_CLAIMS:
            break
    if not claims:
        return None
    return DemoEvidenceExtraction(
        source_type="market",
        title=f"Uploaded tabular research - {Path(filename).stem}"[:255],
        publisher="User-provided tabular source",
        published_at=now,
        claims=tuple(claims),
        instruments=tuple(dict.fromkeys(instruments)),
        parser_name="spi-generic-research-table",
    )


def _csv_rows(content: bytes) -> list[list[object]]:
    decoded: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        return []
    try:
        dialect = csv.Sniffer().sniff(decoded[:16_384], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [list(row) for row in csv.reader(io.StringIO(decoded), dialect)]


def extract_demo_evidence(filename: str, content: bytes) -> DemoEvidenceExtraction | None:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        return extract_research_text(filename, _pdf_text(content))
    if suffix in {".xls", ".xlsx"}:
        return extract_research_rows(filename, _sheet_rows(filename, content))
    if suffix == ".csv":
        return extract_research_rows(filename, _csv_rows(content))
    return None
