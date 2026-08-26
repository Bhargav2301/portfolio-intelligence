from __future__ import annotations

import csv
import hashlib
import io
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any


PDF_SIGNATURE = b"%PDF-"
OLE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
ZIP_SIGNATURES = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")
ALLOWED_SUFFIXES = {".pdf", ".xls", ".xlsx", ".csv"}
UNSUPPORTED_SUFFIXES = {".xlsm", ".xlsb", ".ods"}
PDF_ACTIVE_MARKERS = (b"/JavaScript", b"/JS", b"/Launch", b"/EmbeddedFile")
MAX_PDF_PAGES = 200
MAX_WORKBOOK_SHEETS = 50
MAX_WORKBOOK_ROWS = 500_000
MAX_ZIP_UNCOMPRESSED = 250 * 1024 * 1024
MAX_COMPRESSION_RATIO = 200


class UnsafeFileError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class ParseSummary:
    detected_type: str
    state: str
    authority_level: str
    structure: dict[str, Any]
    warnings: list[str]
    sha256: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def authority_for(source_role: str) -> str:
    return {
        "brokerage_ledger": "ledger_candidate",
        "broker_statement": "snapshot_candidate",
        "pms_statement": "snapshot_candidate",
        "research": "evidence_only",
        "manual": "manual_candidate",
    }.get(source_role, "unknown")


def validate_and_summarize(
    filename: str,
    content: bytes,
    source_role: str,
    declared_type: str | None = None,
) -> ParseSummary:
    del declared_type
    suffix = Path(filename).suffix.lower()
    if suffix in UNSUPPORTED_SUFFIXES:
        raise UnsafeFileError(
            "FILE_TYPE_UNSUPPORTED",
            f"{suffix} is not supported because it may contain active or unsupported content.",
        )
    if suffix not in ALLOWED_SUFFIXES:
        raise UnsafeFileError("FILE_TYPE_UNSUPPORTED", "Use PDF, XLS, XLSX, or CSV.")
    if not content:
        raise UnsafeFileError("EMPTY_FILE", "The uploaded file is empty.")

    sha256 = hashlib.sha256(content).hexdigest()
    authority = authority_for(source_role)
    if suffix == ".pdf":
        structure, warnings, state = _summarize_pdf(content)
        detected = "application/pdf"
    elif suffix == ".xlsx":
        _validate_zip_container(content)
        structure, warnings = _summarize_xlsx(content)
        state = "review_required"
        detected = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif suffix == ".xls":
        if not content.startswith(OLE_SIGNATURE):
            raise UnsafeFileError("SIGNATURE_MISMATCH", "The file is not a valid legacy XLS file.")
        structure, warnings = _summarize_xls(content)
        state = "review_required"
        detected = "application/vnd.ms-excel"
    else:
        structure, warnings = _summarize_csv(content)
        state = "review_required"
        detected = "text/csv"

    return ParseSummary(
        detected_type=detected,
        state=state,
        authority_level=authority,
        structure=structure,
        warnings=warnings,
        sha256=sha256,
    )


def _summarize_pdf(content: bytes) -> tuple[dict[str, Any], list[str], str]:
    if not content.startswith(PDF_SIGNATURE):
        raise UnsafeFileError("SIGNATURE_MISMATCH", "The file is not a valid PDF.")
    if any(marker in content for marker in PDF_ACTIVE_MARKERS):
        raise UnsafeFileError(
            "PDF_ACTIVE_CONTENT",
            "PDF JavaScript, launch actions, attachments, or embedded files are not accepted.",
        )
    from pypdf import PdfReader

    try:
        reader = PdfReader(io.BytesIO(content), strict=True)
    except Exception as error:
        raise UnsafeFileError("PDF_PARSE_FAILED", "The PDF structure could not be read safely.") from error
    if reader.is_encrypted:
        return (
            {"pages": None, "encrypted": True},
            ["A password is required and must be used only for this parsing job."],
            "password_required",
        )
    page_count = len(reader.pages)
    if page_count > MAX_PDF_PAGES:
        raise UnsafeFileError("PDF_PAGE_LIMIT", f"PDF exceeds the {MAX_PDF_PAGES}-page limit.")
    text_pages = 0
    page_character_counts: list[int] = []
    for page in reader.pages:
        try:
            extracted = page.extract_text() or ""
        except Exception:
            extracted = ""
        count = len(extracted.strip())
        page_character_counts.append(count)
        if count >= 20:
            text_pages += 1
    scanned_pages = page_count - text_pages
    warnings: list[str] = []
    if scanned_pages:
        warnings.append(
            f"{scanned_pages} page(s) appear scanned or contain little native text; OCR review is required."
        )
    return (
        {
            "pages": page_count,
            "native_text_pages": text_pages,
            "scanned_or_low_text_pages": scanned_pages,
            "page_character_counts": page_character_counts,
            "encrypted": False,
        },
        warnings,
        "review_required",
    )


def _validate_zip_container(content: bytes) -> None:
    if not content.startswith(ZIP_SIGNATURES):
        raise UnsafeFileError("SIGNATURE_MISMATCH", "The file is not a valid XLSX container.")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            total_compressed = sum(max(item.compress_size, 1) for item in infos)
            total_uncompressed = sum(item.file_size for item in infos)
            names = {item.filename.lower() for item in infos}
    except zipfile.BadZipFile as error:
        raise UnsafeFileError("XLSX_PARSE_FAILED", "The XLSX container is damaged.") from error
    if total_uncompressed > MAX_ZIP_UNCOMPRESSED:
        raise UnsafeFileError("DECOMPRESSION_LIMIT", "Workbook expands beyond the safe limit.")
    if total_uncompressed / max(total_compressed, 1) > MAX_COMPRESSION_RATIO:
        raise UnsafeFileError("COMPRESSION_RATIO_LIMIT", "Workbook compression ratio is unsafe.")
    if any(name.endswith("vbaproject.bin") for name in names):
        raise UnsafeFileError("WORKBOOK_MACRO", "Macro-enabled workbooks are not accepted.")


def _summarize_xlsx(content: bytes) -> tuple[dict[str, Any], list[str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as error:
        raise UnsafeFileError("XLSX_PARSE_FAILED", "The workbook could not be read safely.") from error
    if len(workbook.sheetnames) > MAX_WORKBOOK_SHEETS:
        workbook.close()
        raise UnsafeFileError(
            "WORKBOOK_SHEET_LIMIT",
            f"Workbook exceeds the {MAX_WORKBOOK_SHEETS}-sheet limit.",
        )
    sheets: list[dict[str, Any]] = []
    total_rows = 0
    for sheet in workbook.worksheets:
        row_count = sheet.max_row or 0
        total_rows += row_count
        if total_rows > MAX_WORKBOOK_ROWS:
            workbook.close()
            raise UnsafeFileError(
                "WORKBOOK_ROW_LIMIT",
                f"Workbook exceeds the {MAX_WORKBOOK_ROWS:,}-row limit.",
            )
        preview: list[list[str | None]] = []
        for row in sheet.iter_rows(min_row=1, max_row=min(row_count, 5), values_only=True):
            preview.append([None if value is None else str(value)[:200] for value in row[:30]])
        sheets.append(
            {
                "name": sheet.title,
                "rows": row_count,
                "columns": sheet.max_column or 0,
                "state": sheet.sheet_state,
                "preview": preview,
            }
        )
    workbook.close()
    warnings = [
        "Formulas are not executed; only stored cell values are considered.",
        "Detected rows remain candidates until reconciliation and human publication.",
    ]
    return {"sheets": sheets, "total_rows": total_rows}, warnings


def _summarize_xls(content: bytes) -> tuple[dict[str, Any], list[str]]:
    try:
        import xlrd
    except ImportError as error:
        raise UnsafeFileError(
            "XLS_PARSER_UNAVAILABLE",
            "Legacy XLS support requires the isolated xlrd parser.",
        ) from error
    try:
        workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as error:
        raise UnsafeFileError("XLS_PARSE_FAILED", "The legacy workbook could not be read.") from error
    if workbook.nsheets > MAX_WORKBOOK_SHEETS:
        workbook.release_resources()
        raise UnsafeFileError("WORKBOOK_SHEET_LIMIT", "Workbook has too many sheets.")
    sheets: list[dict[str, Any]] = []
    total_rows = 0
    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)
        total_rows += sheet.nrows
        if total_rows > MAX_WORKBOOK_ROWS:
            workbook.release_resources()
            raise UnsafeFileError("WORKBOOK_ROW_LIMIT", "Workbook has too many rows.")
        preview = [
            [str(value)[:200] for value in sheet.row_values(row_index, 0, min(sheet.ncols, 30))]
            for row_index in range(min(sheet.nrows, 5))
        ]
        sheets.append(
            {"name": sheet_name, "rows": sheet.nrows, "columns": sheet.ncols, "preview": preview}
        )
    workbook.release_resources()
    return (
        {"sheets": sheets, "total_rows": total_rows},
        [
            "Legacy XLS was parsed without executing formulas or macros.",
            "Detected rows remain candidates until reconciliation and human publication.",
        ],
    )


def _summarize_csv(content: bytes) -> tuple[dict[str, Any], list[str]]:
    if b"\x00" in content:
        raise UnsafeFileError("CSV_BINARY_CONTENT", "The CSV contains binary null bytes.")
    decoded: str | None = None
    encoding = ""
    for candidate in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            decoded = content.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise UnsafeFileError("CSV_ENCODING_UNSUPPORTED", "The CSV text encoding is unsupported.")
    sample = decoded[:16_384]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(decoded), dialect)
    preview: list[list[str]] = []
    row_count = 0
    max_columns = 0
    for row in reader:
        row_count += 1
        max_columns = max(max_columns, len(row))
        if row_count <= 5:
            preview.append([value[:200] for value in row[:30]])
        if row_count > MAX_WORKBOOK_ROWS:
            raise UnsafeFileError("CSV_ROW_LIMIT", "CSV has too many rows.")
    return (
        {
            "encoding": encoding,
            "delimiter": dialect.delimiter,
            "rows": row_count,
            "columns": max_columns,
            "preview": preview,
        },
        ["Detected rows remain candidates until reconciliation and human publication."],
    )
